#!/usr/bin/env python3
"""
Turbolink 活动数据拉取脚本
单表纵向堆叠：每个活动的数据依次排列，一行=一个活动+任务
数据统计使用活动本身的起止时间

============================================================
指标计算公式说明
============================================================

1. UV（独立访客数）
   来源：活动列表接口的 visitor_num 字段
   过滤：仅保留 UV > UV_THRESHOLD 的活动

2. 活动参与率
   公式：点击活动主按钮人数 / UV
   说明：激励养成活动使用"领养人数"作为点击数

3. 任务完成率
   公式：完成任务人数 / UV

4. 次留（次日留存率）
   公式：活动开始后第2天仍活跃的用户数 / 活动当天的用户数
   来源：留存接口 day_1 字段

5. 7日留（7日留存率）
   公式：活动开始后第8天仍活跃的用户数 / 活动当天的用户数
   来源：留存接口 day_7 字段

6. 回收金币（仅金币大派送活动）
   公式：用户投放金币总数 - 用户领取金币总数
   说明：按天累加，统计范围为活动开始至 COIN_CUTOFF_DATE

7. 金币回收比例（仅金币大派送活动）
   公式：回收金币数 / 用户投放金币总数
   说明：反映平台从用户侧净回收的金币占用户总投入的比例

============================================================
使用说明
============================================================
1. 在配置区填入 TOKEN（从 Turbolink 后台获取）
2. 填入 PROJECT_ID（项目 ID）
3. 可选：调整 UV_THRESHOLD、COIN_CUTOFF_DATE 等参数
4. 运行：python3 turbolink_fetch_client.py
5. 输出：当前目录下的 Excel 文件
"""

import argparse
import requests
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ============ 配置区（只需改这里）============

# 【必填】Token：登录 Turbolink 后台后，从浏览器请求头中复制 Authorization 值
TOKEN = "Bearer 你的token粘贴在这里"

# 【必填】项目 ID：Turbolink 后台 URL 中的 pjid 参数值
PROJECT_ID = "d1qcfa01bc5m7ka0t450"

# 活动搜索时间范围
SEARCH_START = "2025/01/01 00:00"
SEARCH_END = "2099/12/31 23:59"

# UV 过滤阈值：仅拉取 UV 大于此值的活动
UV_THRESHOLD = 15

# 金币回收统计截止日期（包含该天）
COIN_CUTOFF_DATE = "2026-06-11"

# 输出文件名
OUTPUT_FILE = "活动数据分析.xlsx"

# ============ 常量 ============
BASE_URL = "https://api.branchcn.com"
SLASH = "/"

LIST_API_PARAMS = {
    "page": 1,
    "per_page": 50,
    "mate_id": PROJECT_ID,
    "project_id": PROJECT_ID,
    "start": SEARCH_START,
    "end": SEARCH_END,
    "targets[]": [1, 2, 3],
}

HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "en,zh-CN;q=0.9,zh;q=0.8",
    "authorization": TOKEN,
    "origin": "https://dashboard.turbolink.cc",
    "referer": "https://dashboard.turbolink.cc/",
    "sec-ch-ua": '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"macOS"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "cross-site",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
    "web-set": f"lang=zh-cn;pjid={PROJECT_ID}",
}


# ============ 带重试的 Session ============
def make_session():
    s = requests.Session()
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[500, 502, 503, 504],
                    allowed_methods=["GET", "POST"])
    adapter = HTTPAdapter(max_retries=retries)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s

SESSION = make_session()


def safe_request(method, url, **kwargs):
    for attempt in range(5):
        try:
            if method == "get":
                resp = SESSION.get(url, **kwargs)
            else:
                resp = SESSION.post(url, **kwargs)
            resp.raise_for_status()
            return resp
        except Exception as e:
            if attempt < 4:
                wait = 2 ** attempt
                print(f"    请求失败 ({e})，{wait}s 后重试...")
                time.sleep(wait)
            else:
                raise


# ============ 数据拉取 ============

def fetch_fission_type_map():
    """获取活动类型 mark -> 中文名 映射"""
    url = f"{BASE_URL}/admin/fission/list"
    resp = safe_request("get", url, headers=HEADERS)
    mapping = {}
    for item in resp.json().get("data", {}).get("list", []):
        mapping[item["mark"]] = item["title"]
    return mapping


def fetch_campaign_list():
    """拉取所有活动（自动翻页）"""
    url = f"{BASE_URL}/admin/fission-campaign/list"
    all_campaigns = []
    page = 1
    while True:
        params = {**LIST_API_PARAMS, "page": page}
        targets = params.pop("targets[]")
        query_params = list(params.items()) + [("targets[]", t) for t in targets]
        resp = safe_request("get", url, headers=HEADERS, params=query_params)
        data = resp.json().get("data", {})
        items = data.get("list", [])
        if not items:
            break
        all_campaigns.extend(items)
        total = data.get("total", 0)
        print(f"  活动列表第 {page} 页: 获取 {len(items)} 条 (累计 {len(all_campaigns)}/{total})")
        if len(all_campaigns) >= total:
            break
        page += 1
    return all_campaigns


def filter_campaigns(campaigns, type_map, type_filter=None, start_date=None):
    """过滤 UV > 阈值的活动"""
    # 构建反向映射：中文名 -> fission_mark
    reverse_map = {v: k for k, v in type_map.items()}

    filtered = []
    print(f"\n过滤条件: UV > {UV_THRESHOLD}")
    if type_filter:
        print(f"  活动类型: {', '.join(type_filter)}")
    if start_date:
        print(f"  活动开始 >= {start_date}")
    print(f"{'活动类型':<16} {'活动标题':<24} {'UV':>8} {'状态':>6}")
    print("-" * 60)
    for c in campaigns:
        name = c.get("title", "未知活动")
        fmark = c.get("fission_mark", "")
        act_type = type_map.get(fmark, fmark)
        rid = c.get("id")
        start = c.get("start_date", "")[:10]
        end = c.get("end_date", "")[:10]
        uv = c.get("visitor_num", 0) or 0
        if not rid:
            continue

        # 活动类型筛选
        if type_filter:
            match = False
            for tf in type_filter:
                if tf == fmark or tf == act_type or reverse_map.get(tf) == fmark:
                    match = True
                    break
            if not match:
                continue

        # 活动开始日期筛选
        act_start = start.replace("/", "-")
        act_end = end.replace("/", "-")
        if start_date and act_start < start_date:
            continue

        status = "保留" if uv > UV_THRESHOLD else "跳过"
        print(f"  {act_type:<14} {name:<24} {uv:>8} {status:>6}")
        if uv > UV_THRESHOLD:
            filtered.append({
                "company": "",
                "activity": act_type,
                "fission_mark": fmark,
                "start": act_start,
                "end": act_end,
                "report_id": str(rid),
                "uv": uv,
            })
    return filtered


def fetch_uv(report_id, start, end):
    """获取 UV（独立访客数）"""
    url = f"{BASE_URL}/fbi/report/overview"
    params = {"id": report_id, "start": start, "end": end, "t_offset": 8}
    resp = safe_request("get", url, headers=HEADERS, params=params)
    for item in resp.json().get("data", {}).get("list", []):
        if item.get("mark") == "uv":
            return item.get("total", 0)
    return 0


def fetch_metric(report_id, metric_type, dems, start, end):
    """获取通用指标（点击、分享、下载、注册等）"""
    url = f"{BASE_URL}/fbi/report/custom"
    payload = {"id": report_id, "type": metric_type, "start": start, "end": end, "dems": dems, "t_offset": 8}
    resp = safe_request("post", url, headers=HEADERS, json=payload)
    for row in resp.json().get("data", {}).get("list", []):
        if all(t.get("name") == "all" for t in row.get("title", [])):
            return row.get("user_num", 0)
    return 0


def fetch_task_data(report_id, start, end):
    """获取任务数据（任务名称、完成人数、完成次数）"""
    url = f"{BASE_URL}/fbi/report/custom"
    payload = {"id": report_id, "type": 11, "start": start, "end": end, "dems": ["cond_id"], "t_offset": 8}
    resp = safe_request("post", url, headers=HEADERS, json=payload)
    tasks = []
    for row in resp.json().get("data", {}).get("list", []):
        if all(t.get("name") == "all" for t in row.get("title", [])):
            continue
        tasks.append({
            "task_name": row["title"][0]["name"] if row.get("title") else "未知任务",
            "user_num": row.get("user_num", 0),
            "num": row.get("num", 0),
        })
    return tasks


def fetch_sign_in_tasks(report_id, start, end):
    """签到类活动：签到参与人数和次数"""
    url = f"{BASE_URL}/fbi/report/custom"
    payload = {"id": report_id, "type": 14, "start": start, "end": end, "dems": ["level_mark"], "t_offset": 8}
    resp = safe_request("post", url, headers=HEADERS, json=payload)
    for row in resp.json().get("data", {}).get("list", []):
        if all(t.get("name") == "all" for t in row.get("title", [])):
            return [{"task_name": "签到参与", "user_num": row.get("user_num", 0), "num": row.get("num", 0)}]
    return [{"task_name": "签到参与", "user_num": 0, "num": 0}]


def fetch_grow_data(report_id, start, end):
    """激励养成活动：领养人数和养成人数"""
    url = f"{BASE_URL}/fbi/report/custom"
    result = {"click": 0, "adopt": 0}
    for mark_type, key in [(16, "click"), (30, "adopt")]:
        payload = {"id": report_id, "type": mark_type, "start": start, "end": end, "dems": ["level_mark"], "t_offset": 8}
        resp = safe_request("post", url, headers=HEADERS, json=payload)
        for row in resp.json().get("data", {}).get("list", []):
            titles = row.get("title", [])
            if all(t.get("name") == "all" for t in titles):
                result[key] = row.get("user_num", 0)
                break
    return result


def fetch_stay_data(report_id, start, end):
    """留存数据（次留、7日留）"""
    url = f"{BASE_URL}/fbi/report/stay"
    payload = {"id": report_id, "start": start, "end": end}
    resp = safe_request("post", url, headers=HEADERS, json=payload)
    target = (datetime.strptime(start, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    for row in resp.json().get("data", {}).get("list", []):
        if row.get("date") == target:
            def parse_pct(val):
                val = str(val).replace("%", "").strip()
                return float(val) / 100 if val and val != "-" else 0
            return {
                "day1": parse_pct(row.get("day_1", "0%")),
                "day7": parse_pct(row.get("day_7", "0%")),
            }
    return {"day1": 0, "day7": 0}


def fetch_coin_recovery(report_id, start, end):
    """金币大派送活动：回收金币和用户投放总数
    公式：回收金币 = 用户投放金币总数 - 用户领取金币总数
    统计范围：活动开始至 COIN_CUTOFF_DATE（含）
    """
    url = f"{BASE_URL}/admin/fission-stat/reward-detail"
    params = {"id": report_id}
    resp = safe_request("get", url, headers=HEADERS, params=params)
    data = resp.json().get("data", {})
    pools = data.get("pools", [])
    cutoff = datetime.strptime(COIN_CUTOFF_DATE, "%Y-%m-%d")
    total_user = 0
    total_get = 0
    for p in pools:
        date_str = p["title"].split("(")[0]
        pool_date = datetime.strptime(date_str, "%Y-%m-%d")
        if pool_date > cutoff:
            continue
        total_user += p.get("user_coin", 0)
        total_get += p.get("get_coin", 0)
    recovery = total_user - total_get
    # 投放金币人数(type=38)、领取金币人数(type=39)
    put_num = fetch_metric(report_id, 38, ["platform"], start, end)
    get_num = fetch_metric(report_id, 39, ["platform"], start, end)
    return recovery, total_user, put_num, get_num


def fetch_match_data(report_id):
    """赛事竞猜活动：获取每场比赛数据"""
    url = f"{BASE_URL}/admin/fission-stat/reward-detail"
    params = {"id": report_id}
    resp = safe_request("get", url, headers=HEADERS, params=params)
    data = resp.json().get("data", {})
    schedules = data.get("schedules") or []
    matches = []
    for s in schedules:
        teams = s.get("teams", [])
        votes = s.get("votes", [])
        team_a = teams[0]["name"] if len(teams) > 0 else ""
        team_b = teams[1]["name"] if len(teams) > 1 else ""
        a_votes = votes[0].get("vote_user_num", 0) if len(votes) > 0 else 0
        b_votes = votes[1].get("vote_user_num", 0) if len(votes) > 1 else 0
        draw_votes = votes[2].get("vote_user_num", 0) if len(votes) > 2 else 0
        total_votes = a_votes + draw_votes + b_votes
        # 胜队判断（只显示A队/B队，不显示具体队名）
        win_type = s.get("win_type", 0)
        win_team_id = s.get("win_team_id", "")
        if win_type == 0:
            winner = "/"
        elif win_type == 2:
            winner = "平局"
        elif win_team_id and len(teams) > 0:
            if win_team_id == teams[0].get("id"):
                winner = "A队"
            elif len(teams) > 1 and win_team_id == teams[1].get("id"):
                winner = "B队"
            else:
                winner = "/"
        else:
            winner = "/"
        reward_num = s.get("reward_num", 0)
        # 领奖率 = 领奖人数 / 胜队投票人数
        if win_type == 1 and win_team_id:
            if len(teams) > 0 and win_team_id == teams[0].get("id"):
                win_votes = a_votes
            elif len(teams) > 1 and win_team_id == teams[1].get("id"):
                win_votes = b_votes
            else:
                win_votes = 0
        elif win_type == 2:
            win_votes = draw_votes
        else:
            win_votes = 0
        reward_rate = reward_num / win_votes if win_votes > 0 else 0
        matches.append({
            "vs_date": s.get("vs_date", ""),
            "team_a": team_a,
            "draw_votes": draw_votes,
            "team_b": team_b,
            "a_votes": a_votes,
            "b_votes": b_votes,
            "total_votes": total_votes,
            "winner": winner,
            "reward_num": reward_num,
            "reward_rate": reward_rate,
        })
    matches.sort(key=lambda x: x["vs_date"])
    return matches


def fetch_vote_distribution(report_id, uv):
    """赛事竞猜：从用户列表接口获取投票分布"""
    url = f"{BASE_URL}/admin/fission-user/new-list"
    all_vote_nums = []
    page = 1
    per_page = 1000
    while True:
        params = {
            "pid_type": 0, "t_offset": 8, "start": "", "end": "",
            "page": page, "per_page": per_page, "id": report_id,
        }
        resp = safe_request("get", url, headers=HEADERS, params=params)
        data = resp.json().get("data", {})
        users = data.get("list", [])
        if not users:
            break
        for u in users:
            vn = u.get("vs", {}).get("vote_num", 0)
            all_vote_nums.append(vn)
        total = data.get("total", 0)
        if len(all_vote_nums) >= total:
            break
        page += 1
    voted = [v for v in all_vote_nums if v > 0]
    if not voted:
        return []
    max_vote = max(voted)
    ranges = [(1, 1)]
    upper = 1
    while upper < max_vote:
        lower = upper + 1
        upper = min(upper + 10, max_vote)
        ranges.append((lower, upper))
        if upper >= max_vote:
            break
    result = []
    total_users = len(voted)
    total_vote_sum = sum(voted)
    result.append({"range": "总计", "user_num": total_users, "total_votes": total_vote_sum})
    for lo, hi in ranges:
        count = sum(1 for v in voted if lo <= v <= hi)
        if count == 0:
            continue
        if lo == hi:
            label = f"{lo}票"
        else:
            label = f">{lo-1}且<={hi}"
        result.append({"range": label, "user_num": count, "total_votes": ""})
    return result


def fetch_activity(act):
    """拉取单个活动的完整数据"""
    rid = act["report_id"]
    s, e = act["start"], act["end"]
    fmark = act.get("fission_mark", "")
    is_sign_in = (fmark == "sign_in")
    is_grow = (fmark == "keep")
    is_coin = (fmark == "coin")
    is_vs = (fmark == "vs")
    stay = fetch_stay_data(rid, s, e)

    if is_sign_in:
        tasks = fetch_sign_in_tasks(rid, s, e)
        click = 0
    else:
        tasks = fetch_task_data(rid, s, e)
        click = fetch_metric(rid, 5, ["channel_id"], s, e)

    grow_click = 0
    grow_adopt = 0
    if is_grow:
        grow = fetch_grow_data(rid, s, e)
        grow_click = grow["click"]
        grow_adopt = grow["adopt"]

    coin_recovery = 0
    coin_user_total = 0
    coin_put_num = 0
    coin_get_num = 0
    if is_coin:
        coin_recovery, coin_user_total, coin_put_num, coin_get_num = fetch_coin_recovery(rid, s, e)

    matches = []
    vote_dist = []
    if is_vs:
        matches = fetch_match_data(rid)
        vote_dist = fetch_vote_distribution(rid, act.get("uv", 0))

    return {
        "company": act.get("company", ""),
        "activity": act["activity"],
        "period": f"{s} ~ {e}",
        "is_grow": is_grow,
        "is_coin": is_coin,
        "is_vs": is_vs,
        "uv": act.get("uv", 0),
        "click": click,
        "grow_click": grow_click,
        "grow_adopt": grow_adopt,
        "tasks": tasks,
        "day1": stay["day1"], "day7": stay["day7"],
        "coin_recovery": coin_recovery,
        "coin_user_total": coin_user_total,
        "coin_put_num": coin_put_num,
        "coin_get_num": coin_get_num,
        "matches": matches,
        "vote_dist": vote_dist,
    }


# ============ Excel ============

def set_cell(ws, row, col, value=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    if fmt:
        cell.number_format = fmt
    return cell


def build_excel(all_data, start_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = f"活动数据_开始>={start_date}" if start_date else "活动数据"

    pct = "0.00%"
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, size=11, color="FFFFFF")

    normal_data = [d for d in all_data if not d.get("is_vs")]
    vs_data = [d for d in all_data if d.get("is_vs")]

    # ============ 普通活动 ============
    # A-D: 基础信息, E-F: 点击参与率, G-J: 任务, K-L: 留存, M-R: 金币
    headers = [
        "公司名称", "活动名称", "活动时间", "UV",                    # A-D
        "点击活动主按钮人数", "活动参与率",                            # E-F
        "任务名称", "完成任务人数", "任务完成率", "完成任务次数",        # G-J
        "次留", "7日留", "回收金币", "金币回收比例",                  # K-N
        "投放金币人数", "投放金币率", "领取金币人数", "领取金币率",  # O-R
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

    row = 2
    for d in normal_data:
        first_row = row
        tasks = d["tasks"]
        is_grow = d.get("is_grow", False)
        has_grow_adopt = is_grow and d.get("grow_adopt", 0) > 0
        num_rows = max(1, len(tasks) + (1 if has_grow_adopt else 0)) if is_grow else max(1, len(tasks))

        for i in range(num_rows):
            is_first = (i == 0)
            is_grow_adopt_row = is_grow and has_grow_adopt and i == 1
            task = tasks[i] if 0 <= i < len(tasks) else None

            if is_first:
                set_cell(ws, row, 1, d["company"])
                set_cell(ws, row, 2, d["activity"])
                set_cell(ws, row, 3, d["period"])
                set_cell(ws, row, 4, d["uv"])
            else:
                for c in range(1, 5):
                    set_cell(ws, row, c)

            if is_first:
                click_val = d["grow_click"] if is_grow else d["click"]
                set_cell(ws, row, 5, click_val if click_val > 0 else SLASH)
                set_cell(ws, row, 6, f'=IF(OR(D{first_row}=0,E{first_row}="/"),"/",E{first_row}/D{first_row})', pct)
            elif is_grow_adopt_row:
                set_cell(ws, row, 5, d["grow_adopt"])
                set_cell(ws, row, 6, f'=IF(OR(D{first_row}=0,E{row}=0),"/",E{row}/D{first_row})', pct)
            else:
                set_cell(ws, row, 5)
                set_cell(ws, row, 6)

            if task:
                set_cell(ws, row, 7, task["task_name"])
                set_cell(ws, row, 8, task["user_num"] if task["user_num"] > 0 else SLASH)
                set_cell(ws, row, 9, f'=IF(OR($D${first_row}=0,H{row}="/"),"/",H{row}/$D${first_row})', pct)
                set_cell(ws, row, 10, task["num"] if task["num"] > 0 else SLASH)
            else:
                set_cell(ws, row, 7, SLASH)
                set_cell(ws, row, 8, SLASH)
                set_cell(ws, row, 9)
                set_cell(ws, row, 10, SLASH)

            if is_first:
                set_cell(ws, row, 11, d["day1"] if d["day1"] > 0 else SLASH, pct)
                set_cell(ws, row, 12, d["day7"] if d["day7"] > 0 else SLASH, pct)
            else:
                set_cell(ws, row, 11)
                set_cell(ws, row, 12)

            if is_first and d.get("is_coin"):
                set_cell(ws, row, 13, d["coin_recovery"] if d["coin_recovery"] > 0 else SLASH)
                ratio = d["coin_recovery"] / d["coin_user_total"] if d["coin_user_total"] > 0 else 0
                set_cell(ws, row, 14, ratio if ratio > 0 else SLASH, pct)
            else:
                set_cell(ws, row, 13)
                set_cell(ws, row, 14)

            if is_first and d.get("is_coin"):
                put_num = d.get("coin_put_num", 0)
                get_num = d.get("coin_get_num", 0)
                set_cell(ws, row, 15, put_num if put_num > 0 else SLASH)
                set_cell(ws, row, 16, f'=IF(OR($D${first_row}=0,O{row}="/"),"/",O{row}/$D${first_row})', pct)
                set_cell(ws, row, 17, get_num if get_num > 0 else SLASH)
                set_cell(ws, row, 18, f'=IF(OR($D${first_row}=0,Q{row}="/"),"/",Q{row}/$D${first_row})', pct)
            else:
                for c in range(15, 19):
                    set_cell(ws, row, c)

            row += 1

    # ============ 赛事竞猜活动（在普通活动之后）============
    if vs_data:
        # 空一行分隔
        row += 1
        # vs 表头：A-D 基础信息, E-F 留空, G-J 任务, K-N 投票分布
        vs_headers = [
            "公司名称", "活动名称", "活动时间", "UV",                    # A-D
            "", "",                                                      # E-F 留空
            "任务名称", "完成任务人数", "任务完成率", "完成任务次数",        # G-J
            "投票范围", "投票人数", "参与率", "投票总次数",              # K-N
        ]
        for i, h in enumerate(vs_headers, 1):
            if not h:
                continue
            c = ws.cell(row=row, column=i, value=h)
            c.font = hfont
            c.fill = hfill
            c.alignment = Alignment(horizontal="center")
            c.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
        # 留空列也加边框
        for i in [5, 6]:
            c = ws.cell(row=row, column=i)
            c.font = hfont
            c.fill = hfill
            c.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
        row += 1

        for d in vs_data:
            first_row = row
            tasks = d["tasks"]
            vote_dist = d.get("vote_dist", [])
            uv = d["uv"]
            # 任务和投票分布按行排列
            num_rows = max(len(tasks), len(vote_dist), 1)
            for i in range(num_rows):
                is_first = (i == 0)
                # A-D: 活动信息（仅第一行）
                if is_first:
                    set_cell(ws, row, 1, d["company"])
                    set_cell(ws, row, 2, d["activity"])
                    set_cell(ws, row, 3, d["period"])
                    set_cell(ws, row, 4, uv)
                else:
                    for c in range(1, 5):
                        set_cell(ws, row, c)
                # E-F: 留空
                set_cell(ws, row, 5)
                set_cell(ws, row, 6)
                # G-J: 任务
                if i < len(tasks):
                    t = tasks[i]
                    set_cell(ws, row, 7, t["task_name"])
                    set_cell(ws, row, 8, t["user_num"] if t["user_num"] > 0 else SLASH)
                    set_cell(ws, row, 9, f'=IF(OR($D${first_row}=0,H{row}="/"),"/",H{row}/$D${first_row})', pct)
                    set_cell(ws, row, 10, t["num"] if t["num"] > 0 else SLASH)
                else:
                    for c in range(7, 11):
                        set_cell(ws, row, c)
                # K-N: 投票分布
                if i < len(vote_dist):
                    vd = vote_dist[i]
                    set_cell(ws, row, 11, vd["range"])
                    set_cell(ws, row, 12, vd["user_num"] if vd["user_num"] > 0 else SLASH)
                    set_cell(ws, row, 13, f'=IF(OR($D${first_row}=0,L{row}="/"),"/",L{row}/$D${first_row})', pct)
                    set_cell(ws, row, 14, vd["total_votes"] if vd["total_votes"] else SLASH)
                else:
                    for c in range(11, 15):
                        set_cell(ws, row, c)
                row += 1

    # 列宽
    widths = [14, 14, 24, 8, 20, 10, 40, 12, 10, 12, 14, 12, 10, 12, 14, 12, 14, 12]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    wb.save(OUTPUT_FILE)


# ============ 主流程 ============

def main():
    parser = argparse.ArgumentParser(description="Turbolink 活动数据拉取")
    parser.add_argument("-t", "--activity-type", action="append",
                        help="活动类型（中文名如 '赛事竞猜' 或 fission_mark 如 'vs'），可多次指定")
    parser.add_argument("-s", "--start-date",
                        help="筛选活动开始日期 >= 此值（格式: 2025-01-01）")
    args = parser.parse_args()

    print(f"活动搜索范围: {SEARCH_START} ~ {SEARCH_END}")

    print("正在获取活动类型映射...")
    type_map = fetch_fission_type_map()
    print(f"  共 {len(type_map)} 种活动类型")

    print("正在拉取活动列表...")
    campaigns = fetch_campaign_list()
    print(f"共获取 {len(campaigns)} 个活动")

    activities = filter_campaigns(campaigns, type_map,
                                  type_filter=args.activity_type,
                                  start_date=args.start_date)
    print(f"\n符合条件的活动: {len(activities)} 个")

    if not activities:
        print("没有符合条件的活动，退出")
        return

    print("\n开始拉取详细数据...")
    all_data = []
    for act in activities:
        print(f"\n拉取: {act['activity']} ({act['start']} ~ {act['end']})")
        data = fetch_activity(act)
        all_data.append(data)
        if data.get("is_vs"):
            print(f"  UV: {data['uv']} | 任务: {len(data['tasks'])} 个 | 投票分布: {len(data.get('vote_dist', []))} 段")
        else:
            print(f"  UV: {data['uv']} | 点击: {data['click']} | 任务: {len(data['tasks'])} 个")

    build_excel(all_data, start_date=args.start_date)
    print(f"\nExcel 已保存: {OUTPUT_FILE} ({len(all_data)} 个活动)")


if __name__ == "__main__":
    main()
